import os
import time
import openpyxl as xl
from passporteye import read_mrz
import csv


class Mrz:
    def __init__(self, filename):
        self.filename = filename
        self.mrz = read_mrz(self.filename)
        self.good = self.mrz is not None
        if self.good:
            self.mrz = self.mrz.to_dict()
            self.mrz = {k: str(v) for k, v in self.mrz.items()}
            print(f"**Saving {self.mrz['names'].title()} {self.mrz['surname'].title()}'s passport.**")
            self.good = True
            self.num_filler = '0' * len(self.mrz['number'].replace('<', ''))
            self.per_filler = '0' * len(self.mrz['personal_number'].repalce('<', ''))
        else:
            print(f"**Skipping {os.path.split(self.filename)[1]} because no MRZ was found.**")
            self.good = False

    def format_mrz(self, y2kprefix):
        # remove all occurrences of <
        self.mrz = {k: v.replace('<', '') for k, v in self.mrz.items()}
        # arrange dates to DD/MM/YYYY instead of YYMMDD
        birth_date = self.mrz['date_of_birth']
        expiration_date = self.mrz['expiration_date']
        byear = time.strftime("%y", time.strptime(birth_date, "%y%m%d"))
        eyear = time.strftime("%y", time.strptime(expiration_date, "%y%m%d"))
        byear = f'20{byear}' if int(byear) <= int(y2kprefix) else f'19{byear}'
        eyear = f'20{eyear}' if int(eyear) <= int(y2kprefix) else f'19{eyear}'
        self.mrz['date_of_birth'] = time.strftime(f"%d/%m/{byear}", time.strptime(birth_date, "%y%m%d"))
        self.mrz['expiration_date'] = time.strftime(f"%d/%m/{eyear}", time.strptime(expiration_date, "%y%m%d"))
        if all(c in '0123456789' for c in self.mrz['number']) and self.mrz['number'][0] != '0':
            self.mrz['number'] = int(self.mrz['number'])
        # format ID:
        if self.mrz['mrz_type'] == 'TD3':
            if all(c in '0123456789' for c in self.mrz['personal_number']):
                self.mrz['personal_number'] = int(self.mrz['personal_number'])

    def save_to_xlsx(self, workbook, output_path):
        sheet = workbook.active
        row = sheet.max_row
        sheet.cell(row + 1, 1).value = row
        sheet.cell(row + 1, 2).value = self.mrz['country']
        sheet.cell(row + 1, 3).value = self.mrz['names']
        sheet.cell(row + 1, 4).value = self.mrz['surname']
        sheet.cell(row + 1, 5).value = self.mrz['number']
        sheet.cell(row + 1, 5).number_format = self.num_filler
        sheet.cell(row + 1, 6).value = self.mrz['nationality']
        sheet.cell(row + 1, 7).value = self.mrz['date_of_birth']
        sheet.cell(row + 1, 8).value = self.mrz['sex']
        sheet.cell(row + 1, 9).value = self.mrz['expiration_date']
        if self.mrz['mrz_type'] == 'TD3':
            sheet.cell(row + 1, 10).value = self.mrz['personal_number']
            sheet.cell(row + 1, 10).number_format = self.per_filler

        workbook.save(output_path)
        print(f"Saved {self.mrz['names'].title()} {self.mrz['surname'].title()}'s passport.")

    def save_to_csv(self, output_path):
        with open(output_path, 'r') as f:
            csv_reader = csv.reader(f)
            row = sum(1 for _ in csv_reader)

        data = [row, self.mrz['country'], self.mrz['names'], self.mrz['surname'], self.mrz['number'],
                self.mrz['nationality'], self.mrz['date_of_birth'], self.mrz['sex'], self.mrz['expiration_date']]
        if self.mrz['mrz_type'] == 'TD3':
            data.append(self.mrz['personal_number'])

        with open('test.csv', 'a', newline="") as f:
            csv_writer = csv.writer(f)
            csv_writer.writerow(data)
