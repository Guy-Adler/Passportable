import eel
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory
from passportable import Mrz
import openpyxl as xl
import time
import pandas as pd
import winshell
import traceback


eel.init('web')


@eel.expose
def ask_folder():
    """ Ask the user to select a folder
     :return the folder path, or None.
     """
    try:
        root = Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        folder = askdirectory(parent=root)
        root.update()
        return {'folder': folder} if folder != "" else {}

    except Exception:
        trace = traceback.format_exc().split('\n')[:-1]  # Last line is always empty
        lines = [[int((len(l) - len(l.lstrip(' '))) / 2), l] for l in trace[:-1]]
        trace = {trace[-1]: lines}  # trace[-1] is the exception, the rest is traceback.
        return {'exception': trace}


@eel.expose
def ask_file_save_location():
    """ Ask the user where to save a file
     :return the file path, or None
     """
    try:
        root = Tk()
        root.withdraw()
        root.wm_attributes('-topmost', 1)
        file_types = [('Excel File or CSV File', '*.xlsx *.csv')]
        file_path = askopenfilename(parent=root, filetypes=file_types)
        root.update()

        return {'file_path': file_path} if file_path != '' else None

    except Exception:
        trace = traceback.format_exc().split('\n')[:-1]  # Last line is always empty
        lines = [[int((len(l) - len(l.lstrip(' '))) / 2), l] for l in trace[:-1]]
        trace = {trace[-1]: lines}  # trace[-1] is the exception, the rest is traceback.
        return {'exception': trace}


@eel.expose
def get_files(p):
    """
    Get all of the supported files from a folder
    :param p: path to the folder
    :type p: str
    :return: a list of paths to all files
    :rtype: list[str]
    """
    try:
        images = []
        valid_extensions = ['.jpg', '.jpeg', '.png']
        for f in os.listdir(p):
            if os.path.splitext(f)[1].lower() not in valid_extensions:
                continue
            ftime = time.strftime('%A, %B %d %Y at %H:%M', time.localtime(os.path.getmtime(os.path.join(p, f))))
            fname = os.path.join(p, f).replace('\\', '/')
            images.append([fname, ftime])

        return {'images': images}

    except Exception:
        trace = traceback.format_exc().split('\n')[:-1]  # Last line is always empty
        lines = [[int((len(l) - len(l.lstrip(' '))) / 2), l] for l in trace[:-1]]
        trace = {trace[-1]: lines}  # trace[-1] is the exception, the rest is traceback..
        return {'exception': trace}


@eel.expose
def save_passport(image, xlsx, y2k=30):
    """
    Save the passport to the excel file
    :param image: path to an image
    :type image: str
    :param xlsx: path to xlsx file
    :type xlsx: str
    :param y2k: when should a century change (does 19 mean 2019 or 1919?)
    :type y2k: int
    :return: image name, name of person (if found)
    """
    try:
        wb = xl.load_workbook(xlsx)
        mrz = Mrz(image)
        if mrz.good:
            mrz.format_mrz(y2k)
            if os.path.splitext(xlsx)[1].lower() == '.xlsx':
                mrz.save_to_xlsx(wb, xlsx)
            elif os.path.splitext(xlsx)[1].lower() == '.csv':
                mrz.save_to_csv(xlsx)
            response = {'file': os.path.basename(image),
                        'name': f'{mrz.mrz["names"].title()} {mrz.mrz["surname"].title()}'}
        else:
            response = {'file': os.path.basename(image), 'name': ''}

        return response
    except Exception:
        trace = traceback.format_exc().split('\n')[:-1]  # Last line is always empty
        lines = [[int((len(l) - len(l.lstrip(' '))) / 2), l] for l in trace[:-1]]
        trace = {trace[-1]: lines}  # trace[-1] is the exception, the rest is traceback.
        return {'exception': trace}


@eel.expose
def get_images_folder_content(p):
    """
    Get a 2D list including file name and file type
    :param p: path to folder
    :type p: str
    :return: 2D list including file name and file type
    :rtype: list[list[str, str]]
    """
    try:
        images = []
        valid_extensions = ['.jpg', '.jpeg', '.png']
        for f in os.listdir(p):
            if os.path.splitext(f)[1].lower() not in valid_extensions:
                continue
            ftime = time.strftime('%A, %B %d %Y at %H:%M', time.localtime(os.path.getmtime(os.path.join(p, f))))
            fname = os.path.basename(os.path.join(p, f).replace('\\', '/'))
            images.append([fname, ftime])

        return {'images': images}

    except Exception:
        trace = traceback.format_exc().split('\n')[:-1]  # Last line is always empty
        lines = [[int((len(l) - len(l.lstrip(' '))) / 2), l] for l in trace[:-1]]
        trace = {trace[-1]: lines}  # trace[-1] is the exception, the rest is traceback.
        return {'exception': trace}


@eel.expose
def get_output_file_content(p):
    """"
    Get a 2D list including All of the output file
    :param p: path to file
    :type p: str
    :return: 2D list including the output file
    :rtype: List[List[str]]
    """
    try:
        table = []
        if os.path.splitext(p)[1].lower() == '.xlsx':
            df = pd.read_excel(p, sheet_name=0)
            table = [df.columns.values.tolist()] + df.values.tolist()

        elif os.path.splitext(p)[1].lower() == '.csv':
            df = pd.read_csv(p)
            table = [df.columns.values.tolist()] + df.values.tolist()

        return {'table': table}

    except Exception:
        trace = traceback.format_exc().split('\n')[:-1]  # Last line is always empty
        lines = [[int((len(l) - len(l.lstrip(' '))) / 2), l] for l in trace[:-1]]
        trace = {trace[-1]: lines}  # trace[-1] is the exception, the rest is traceback.
        return {'exception': trace}


@eel.expose
def create_report(images):
    """
    Save report to an xlsx file
    :param images: 2d list of all images
    :return: None
    """
    try:
        wb = xl.Workbook()
        sheet = wb.active
        sheet.cell(1, 1).value = '#'
        sheet.cell(1, 2).value = 'File Name'
        sheet.cell(1, 3).value = 'Last Modified'
        sheet.cell(1, 4).value = 'Skipped'
        row = sheet.max_row
        for image in images:
            sheet.cell(row + 1, 1).value = row
            sheet.cell(row + 1, 2).value = image[0]
            sheet.cell(row + 1, 3).value = image[1]
            sheet.cell(row + 1, 4).value = image[2]
            row += 1

        p = os.path.join(winshell.my_documents(), 'Passportable\\logs')
        if not os.path.exists(p):
            os.makedirs(p)
        name = os.path.join(p, str(time.strftime('%A, %B %d %Y at %H-%M', time.localtime())) + '.xlsx')
        wb.save(filename=name)
        os.startfile(name)
        return {'name': name}

    except Exception:
        trace = traceback.format_exc().split('\n')[:-1]  # Last line is always empty
        lines = [[int((len(l) - len(l.lstrip(' '))) / 2), l] for l in trace[:-1]]
        trace = {trace[-1]: lines}  # trace[-1] is the exception, the rest is traceback.
        return {'exception': trace}


if __name__ == '__main__':
    eel.start('index.html', size=(650, 650))
